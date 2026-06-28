from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.conf import settings
from django.db.models import Max
from .models import Cart, Product, Blank
from users.models import User
import json, time, string, os, hashlib, uuid
from decimal import Decimal
import openpyxl
from openpyxl.styles import Alignment
import requests


def index(request):
    if request.user.is_authenticated:
        return redirect('cart')
    return render(request, 'landing.html', {
        'view': 'index',
    })


def login_view(request):
    if request.user.is_authenticated:
        return redirect('cart')
    if request.method == 'POST':
        login_input = request.POST.get('login')
        password = request.POST.get('password')
        user = authenticate(request, username=login_input, password=password)
        if user is not None:
            login(request, user)
            return redirect('cart')
        else:
            return render(request, 'main.html', {
                'view': 'login',
                'login_error': True,
            })
    return render(request, 'main.html', {'view': 'login'})


def logout_view(request):
    logout(request)
    return redirect('index')


def registration_view(request):
    if request.user.is_authenticated:
        return redirect('cart')
    if request.method == 'POST':
        login_input = request.POST.get('login')
        email = request.POST.get('email')
        password = request.POST.get('password')
        promocode = request.POST.get('promocode', '')

        if User.objects.filter(login=login_input).exists():
            return render(request, 'main.html', {'view': 'registration', 'login_taken': True})
        if User.objects.filter(email=email).exists():
            return render(request, 'main.html', {'view': 'registration', 'email_taken': True})

        user = User.objects.create_user(login=login_input, email=email, password=password)
        login(request, user)

        cart = Cart.objects.create(user=user, name='Список 1')
        request.session['current_cart'] = cart.id

        return redirect('cart')

    promocode = request.GET.get('promocode', '')
    return render(request, 'main.html', {'view': 'registration', 'promocode': promocode})


@login_required
def cart_view(request):
    carts = Cart.objects.filter(user=request.user, deleted=False)
    cart_id = request.GET.get('cart_id')
    if cart_id:
        current_cart = get_object_or_404(Cart, id=cart_id, user=request.user, deleted=False)
        request.session['current_cart'] = current_cart.id
    else:
        current_cart_id = request.session.get('current_cart')
        if current_cart_id:
            try:
                current_cart = Cart.objects.get(id=current_cart_id, user=request.user, deleted=False)
            except Cart.DoesNotExist:
                current_cart = carts.first()
        else:
            current_cart = carts.first()

    if current_cart:
        products = Product.objects.filter(user=request.user, cart=current_cart, deleted=False).order_by('cart_order')
    else:
        products = []

    blanks = Blank.objects.filter(user=request.user)
    balance = request.user.balance
    tariff = request.user.tariff
    total_items = len(products)
    cost = total_items * float(tariff) if tariff else 0
    balance_enough = balance >= cost

    return render(request, 'main.html', {
        'view': 'cart',
        'carts': carts,
        'current_cart': current_cart,
        'products': products,
        'blanks': blanks,
        'total_items': total_items,
        'total_cost': cost,
        'balance_enough': balance_enough,
    })


@login_required
def blanks_view(request):
    blanks = Blank.objects.filter(user=request.user)
    return render(request, 'main.html', {
        'view': 'blanks',
        'blanks': blanks,
    })


@login_required
def configure_view(request, blank_id):
    blank = get_object_or_404(Blank, blank_id=blank_id, user=request.user)
    return render(request, 'main.html', {
        'view': 'configure',
        'blank': blank,
    })


@login_required
def profile_view(request):
    if request.method == 'POST':
        user = request.user
        user.name = request.POST.get('first-name', '')
        user.last_name = request.POST.get('last-name', '')
        user.email = request.POST.get('email', user.email)
        user.gender = int(request.POST.get('gender', 0))
        user.b_day = int(request.POST.get('b_day', 0)) if request.POST.get('b_day') else None
        user.b_month = int(request.POST.get('b_month', 0)) if request.POST.get('b_month') else None
        user.b_year = int(request.POST.get('b_year', 0)) if request.POST.get('b_year') else None

        new_password = request.POST.get('new_password')
        if new_password:
            user.set_password(new_password)

        user.save()
        if new_password:
            login(request, user)
        messages.success(request, 'Профиль обновлен')
        return redirect('profile')

    months = {
        1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель', 5: 'Май', 6: 'Июнь',
        7: 'Июль', 8: 'Август', 9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь',
    }
    years = list(range(1940, 2011))

    return render(request, 'main.html', {
        'view': 'profile',
        'user_obj': request.user,
        'months': months,
        'years': years,
    })


@login_required
def balance_view(request):
    return render(request, 'main.html', {
        'view': 'balance',
    })


def oferta_view(request):
    return render(request, 'main.html', {
        'view': 'oferta',
    })


@csrf_exempt
def check_login(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        login_val = data.get('login', '')
        count = User.objects.filter(login=login_val).count()
        return JsonResponse({'count': count})
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
def check_email(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        email = data.get('email', '')
        count = User.objects.filter(email=email).count()
        return JsonResponse({'count': count})
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
def delete_product(request, product_id):
    product = get_object_or_404(Product, id=product_id, user=request.user)
    product.deleted = True
    product.save()
    return JsonResponse({'status': 'ok'})


@login_required
def delete_blank(request, blank_id):
    blank = get_object_or_404(Blank, blank_id=blank_id, user=request.user)
    blank.delete()
    return JsonResponse({'status': 'ok'})


@login_required
@csrf_exempt
def recount_product(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        product_id = data.get('id')
        new_amount = data.get('newAmount')
        product = get_object_or_404(Product, id=product_id, user=request.user)
        product.amount = new_amount
        product.save()
        return JsonResponse({'status': 'ok'})
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
@csrf_exempt
def update_blank_index(request):
    if request.method == 'POST':
        field = request.POST.get('field')
        value = request.POST.get('value')
        blank_id = request.POST.get('blank_id')
        blank = get_object_or_404(Blank, blank_id=blank_id, user=request.user)

        field_map = {
            'sheet_name': 'target_sheet_name',
            'link': 'link_index',
            'image': 'img_index',
            'color': 'color_index',
            'size': 'size_index',
            'amount': 'amount_index',
            'price': 'price_index',
            'first_row': 'caption_row_index',
        }

        if field in field_map:
            if field == 'first_row':
                value = int(value) - 1
            setattr(blank, field_map[field], value)
            blank.save()
            return JsonResponse({'status': 'ok'})

    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
@csrf_exempt
def manage_cart(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        action = data.get('action')

        if action == 'new_list':
            name = data.get('name', 'Новый список')
            last_id = Cart.objects.aggregate(Max('id'))['id__max'] or 0
            cart = Cart.objects.create(id=last_id + 1, user=request.user, name=name)
            request.session['current_cart'] = cart.id
            return JsonResponse({'status': 'ok', 'id': cart.id})

        if action == 'rename_list':
            cart_id = data.get('cart_id')
            name = data.get('name')
            cart = get_object_or_404(Cart, id=cart_id, user=request.user)
            cart.name = name
            cart.save()
            return JsonResponse({'status': 'ok'})

        if action == 'delete_list':
            cart_id = data.get('cart_id')
            cart = get_object_or_404(Cart, id=cart_id, user=request.user)
            cart.deleted = True
            cart.save()
            remaining = Cart.objects.filter(user=request.user, deleted=False).first()
            if remaining:
                request.session['current_cart'] = remaining.id
            return JsonResponse({'status': 'ok'})

        if action == 'duplicate_list':
            cart_id = data.get('cart_id')
            cart = get_object_or_404(Cart, id=cart_id, user=request.user)
            new_cart = Cart.objects.create(user=request.user, name=f"{cart.name} (Копия)")
            products = Product.objects.filter(cart=cart, deleted=False)
            for p in products:
                Product.objects.create(
                    user=request.user,
                    cart=new_cart,
                    link=p.link,
                    img_url=p.img_url,
                    color=p.color,
                    size=p.size,
                    amount=p.amount,
                    price=p.price,
                    shop_home_page=p.shop_home_page,
                    cart_order=p.cart_order,
                )
            request.session['current_cart'] = new_cart.id
            return JsonResponse({'status': 'ok', 'id': new_cart.id})

        if action == 'move_products':
            cart_id = data.get('destination_cart')
            product_ids = data.get('products', [])
            dest_cart = get_object_or_404(Cart, id=cart_id, user=request.user)
            Product.objects.filter(id__in=product_ids, user=request.user).update(cart=dest_cart)
            return JsonResponse({'status': 'ok'})

        if action == 'delete_products':
            product_ids = data.get('products', [])
            Product.objects.filter(id__in=product_ids, user=request.user).update(deleted=True)
            return JsonResponse({'status': 'ok'})

        if action == 'reorder':
            new_order = data.get('new_order', {})
            for pos, pid in new_order.items():
                Product.objects.filter(id=pid, user=request.user).update(cart_order=int(pos))
            return JsonResponse({'status': 'ok'})

    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
@csrf_exempt
def upload_blank(request):
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        ts_created = int(time.time())
        blank_filename_original = uploaded_file.name
        blank_filename_uniq = f"{ts_created}_{blank_filename_original}"

        blanks_dir = os.path.join(settings.MEDIA_ROOT, 'blanks')
        os.makedirs(blanks_dir, exist_ok=True)
        filepath = os.path.join(blanks_dir, blank_filename_uniq)

        with open(filepath, 'wb') as f:
            for chunk in uploaded_file.chunks():
                f.write(chunk)

        alfa_list = list(string.ascii_uppercase)
        sheet_name = ''
        caption_row_index = 1
        img_index = color_index = size_index = amount_index = price_index = link_index = ''

        try:
            book = openpyxl.load_workbook(filepath)

            def get_points(word):
                points = 0
                keyword_list = ['Фотография', 'Изображение', 'Фото', 'Картинка', 'Размер', 'Величина', 'Разм', 'Ссылка', 'URL', 'URI', 'Цвет', 'Расцветка', 'Количество', 'Кол-во', 'Колич', 'К-во', 'Колво', "шт", 'Цена', 'Стоимость', 'Характеристики', 'Характеристика', 'Хар-ки', 'Характ', 'Описание', 'Комментарий', 'Коментарий', 'Коммент', 'Комент']
                for keyword in keyword_list:
                    if keyword.lower() in word.lower():
                        points += 1
                return points

            def get_rating_list(ws):
                rating = {}
                for j in range(45):
                    rating[j + 1] = 0
                    for alfa in alfa_list:
                        cell = ws[f'{alfa}{j + 1}'].value
                        if cell:
                            rating[j + 1] += get_points(str(cell))
                return rating

            def get_raiting_list_total(rating_list):
                return sum(rating_list.values())

            def get_sheets_points(book):
                sheets_points = {}
                for sn in book.sheetnames:
                    sheets_points[sn] = get_raiting_list_total(get_rating_list(book[sn]))
                    sn_lower = sn.lower()
                    if 'бланк заказа' in sn_lower:
                        sheets_points[sn] += 30
                    elif 'бланк' in sn_lower:
                        sheets_points[sn] += 10
                    elif 'blank' in sn_lower:
                        sheets_points[sn] += 10
                    elif 'форма заказа' in sn_lower:
                        sheets_points[sn] += 30
                return sheets_points

            sheets_points = get_sheets_points(book)
            sheet_name = max(sheets_points, key=sheets_points.get) if sheets_points else book.sheetnames[0]
            ws = book[sheet_name]
            rating_list = get_rating_list(ws)
            caption_row_index = max(rating_list, key=rating_list.get)

            def find_caption_column_index(ws, row_number, group_mark):
                group_mark_lists = {
                    'img': ['Фотография', 'Изображение', 'Фото', 'Картинка'],
                    'size': ['Размер', 'Величина', 'Разм'],
                    'link': ['Ссылка', 'URL', 'URI', 'ссылки', 'ссылк'],
                    'color': ['Цвет', 'Расцветка'],
                    'amount': ['Количество', 'Кол-во', 'Колич', 'К-во', 'Колво', "шт"],
                    'price': ['Цена', 'Стоимость'],
                    'feature': ['Характеристики', 'Характеристика', 'Хар-ки', 'Характ', 'Описание', 'Комментарий', 'Коментарий', 'Коммент', 'Комент']
                }
                name_variants = group_mark_lists[group_mark]
                for alfa in alfa_list:
                    cell_val = str(ws[f'{alfa}{row_number}'].value or '')
                    for keyword in name_variants:
                        if keyword.lower() in cell_val.lower():
                            return alfa
                return ''

            img_index = find_caption_column_index(ws, caption_row_index, 'img')
            color_index = find_caption_column_index(ws, caption_row_index, 'color')
            size_index = find_caption_column_index(ws, caption_row_index, 'size')
            amount_index = find_caption_column_index(ws, caption_row_index, 'amount')
            price_index = find_caption_column_index(ws, caption_row_index, 'price')
            link_index = find_caption_column_index(ws, caption_row_index, 'link')
            feature_index = find_caption_column_index(ws, caption_row_index, 'feature')

            img_index = img_index or find_caption_column_index(ws, caption_row_index + 1, 'img')
            color_index = color_index or find_caption_column_index(ws, caption_row_index + 1, 'color')
            size_index = size_index or find_caption_column_index(ws, caption_row_index + 1, 'size')
            amount_index = amount_index or find_caption_column_index(ws, caption_row_index + 1, 'amount')
            price_index = price_index or find_caption_column_index(ws, caption_row_index + 1, 'price')
            link_index = link_index or find_caption_column_index(ws, caption_row_index + 1, 'link')

            if caption_row_index > 1:
                img_index = img_index or find_caption_column_index(ws, caption_row_index - 1, 'img')
                color_index = color_index or find_caption_column_index(ws, caption_row_index - 1, 'color')
                size_index = size_index or find_caption_column_index(ws, caption_row_index - 1, 'size')
                amount_index = amount_index or find_caption_column_index(ws, caption_row_index - 1, 'amount')
                price_index = price_index or find_caption_column_index(ws, caption_row_index - 1, 'price')
                link_index = link_index or find_caption_column_index(ws, caption_row_index - 1, 'link')

            if not size_index and color_index:
                size_index = color_index
            if not color_index and size_index:
                color_index = size_index
            if not color_index and not size_index:
                color_index = feature_index
                size_index = feature_index

        except Exception:
            pass

        Blank.objects.create(
            user=request.user,
            blank_filename_original=blank_filename_original,
            blank_file=f'blanks/{blank_filename_uniq}',
            target_sheet_name=sheet_name,
            caption_row_index=caption_row_index,
            img_index=img_index,
            color_index=color_index,
            size_index=size_index,
            amount_index=amount_index,
            price_index=price_index,
            link_index=link_index,
        )

        return redirect('blanks')

    return redirect('blanks')


@login_required
@csrf_exempt
def download_filled(request):
    if request.method == 'POST':
        blank_id = request.POST.get('file_blank_name')
        current_list_id = request.POST.get('current_list_id')

        blank = get_object_or_404(Blank, blank_id=blank_id, user=request.user)
        products = Product.objects.filter(user=request.user, cart_id=current_list_id, deleted=False).order_by('cart_order')

        if not products:
            return redirect('cart')

        if not blank.blank_file:
            messages.error(request, 'Файл бланка не найден')
            return redirect('cart')
        filepath = blank.blank_file.path

        if not os.path.exists(filepath):
            messages.error(request, 'Файл бланка не найден')
            return redirect('cart')

        book = openpyxl.load_workbook(filepath)
        sheet = book[blank.target_sheet_name] if blank.target_sheet_name else book.active

        img_index = blank.img_index
        color_index = blank.color_index
        size_index = blank.size_index
        amount_index = blank.amount_index
        price_index = blank.price_index
        link_index = blank.link_index
        start_iter_row = blank.caption_row_index + 1
        alfa_list = list(string.ascii_uppercase)
        max_alfa = max([i for i in [img_index, color_index, size_index, amount_index, price_index, link_index] if i], default='A')
        max_alfa_index = alfa_list.index(max_alfa) if max_alfa in alfa_list else 0

        merged_cells = sheet.merged_cells
        exclude_rows = []
        for i in range(start_iter_row, 200):
            for alfa in alfa_list:
                if alfa_list.index(alfa) <= max_alfa_index:
                    if f'{alfa}{i}' in merged_cells:
                        exclude_rows.append(i)
                        break

        height = 65
        for i in range(start_iter_row, start_iter_row + 30):
            if sheet.row_dimensions[i].height and sheet.row_dimensions[i].height > height:
                height = sheet.row_dimensions[i].height

        height_rows = height
        if img_index:
            sheet.column_dimensions[img_index].width = height_rows / 2.5

        img_dir = os.path.join(settings.MEDIA_ROOT, 'blanks_download_img')
        os.makedirs(img_dir, exist_ok=True)

        iter_rows = start_iter_row

        def add_to_blank(row, exclude_rows):
            nonlocal iter_rows
            while iter_rows in exclude_rows:
                iter_rows += 1
            sheet.row_dimensions[iter_rows].height = height_rows

            img_ulr = row.img_url
            link_val = row.link
            price_val = float(row.price)
            amount_val = float(row.amount)
            color_val = row.color
            size_val = row.size

            img_hash = hashlib.md5((img_ulr + link_val).encode('utf-8')).hexdigest() + ".jpg"
            img_path = os.path.join(img_dir, img_hash)

            if not os.path.exists(img_path) and img_ulr:
                try:
                    headers = {'User-Agent': 'Mozilla/5.0'}
                    img_data = requests.get(img_ulr, headers=headers, timeout=10).content
                    with open(img_path, 'wb') as f:
                        f.write(img_data)
                except Exception:
                    pass

            if img_index and os.path.exists(img_path):
                img = openpyxl.drawing.image.Image(img_path)
                img.width = height_rows * 1.25
                img.height = height_rows * 1.25
                img.anchor = f'{img_index}{iter_rows}'
                sheet.add_image(img)

            if link_index:
                sheet[f'{link_index}{iter_rows}'] = link_val
            if price_index:
                sheet[f'{price_index}{iter_rows}'] = price_val
            if amount_index:
                sheet[f'{amount_index}{iter_rows}'] = amount_val
            if color_index:
                sheet[f'{color_index}{iter_rows}'] = color_val
                sheet[f'{color_index}{iter_rows}'].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            if size_index:
                if color_index == size_index and color_val and size_val:
                    sheet[f'{size_index}{iter_rows}'] = f"{color_val}: {size_val}"
                else:
                    sheet[f'{size_index}{iter_rows}'] = size_val
                sheet[f'{size_index}{iter_rows}'].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

            if price_index:
                sheet[f'{price_index}{iter_rows}'].alignment = Alignment(horizontal='center', vertical='center')
            if amount_index:
                sheet[f'{amount_index}{iter_rows}'].alignment = Alignment(horizontal='center', vertical='center')
            if link_index:
                sheet[f'{link_index}{iter_rows}'].alignment = Alignment(wrapText=True)

            if color_index and img_index and color_index == img_index:
                sheet[f'{color_index}{iter_rows}'].alignment = Alignment(horizontal='right', vertical='center', wrapText=True)
            if size_index and img_index and size_index == img_index:
                sheet[f'{size_index}{iter_rows}'].alignment = Alignment(horizontal='right', vertical='center', wrapText=True)

            iter_rows += 1

        for product in products:
            add_to_blank(product, exclude_rows)

        count = len(products)
        tariff = float(request.user.tariff or 5)
        cost = tariff * count
        new_balance = float(request.user.balance) - cost
        request.user.balance = new_balance
        request.user.save()

        xlsx_filename = blank.blank_filename_original
        if not xlsx_filename.endswith('x'):
            xlsx_filename += 'x'

        download_dir = os.path.join(settings.MEDIA_ROOT, 'blanks_download')
        os.makedirs(download_dir, exist_ok=True)
        output_path = os.path.join(download_dir, xlsx_filename)
        book.save(output_path)

        with open(output_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{xlsx_filename}"'
            return response

    return redirect('cart')
